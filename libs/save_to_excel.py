from datetime import datetime
import os
from PyQt5.QtCore import QThread, pyqtSignal, QSettings
from PyQt5 import QtCore
import pandas as pd
import openpyxl

class SaveExcel(QThread):
    pb_change_visible = QtCore.pyqtSignal()

    def __init__(self,data, file_path):
        super(SaveExcel, self).__init__()
        self.data = data
        self.file_path = file_path

    def run(self):
        wb = openpyxl.Workbook()
        sheet = wb.active
        print(self.data)
        row = 1
        col = 1
        for i in range(len(self.data)):
            sheet.cell(row,1).value = self.data[i]["ubp"] + " (За дату {})".format(self.data[i]["date"])
            row+=1
            for j in range(len(self.data[i]["results"])):
                sheet.cell(row,1).value = self.data[i]["results"][j]["descr"]
                sheet.cell(row,2).value = self.data[i]["results"][j]["result"]
                row+=1
                    
            row +=1 

        row+=1

        sheet.cell(row,1).value = "Протокол проверки сформирован "
        sheet.cell(row,2).value = datetime.now().strftime("%d.%m.%Y")

        sheet.column_dimensions['A'].width = 55
        sheet.column_dimensions['B'].width = 30
        # df = pd.DataFrame(data=self.data)
        # df.style.set_properties(subset=['Описание'], **{'width': '300px'})
        # df.to_excel(self.file_path)
        wb.save(filename=self.file_path)
        self.pb_change_visible.emit()
        
            